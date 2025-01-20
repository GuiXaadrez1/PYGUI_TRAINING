from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
import os

class Formatar:
    """
    Formata as colunas do relatório gerado com base no arquivo XLSX
    presente no servidor.
    
    Args:
        xlsx_servidor (str): Caminho do arquivo no servidor.
    """
    def __init__(self, xlsx_servidor: str):
        self.xlsx_servidor = xlsx_servidor

    def ler_xlsx_DER_MULTAS(self):
        """
        Carrega e exibe informações básicas do arquivo XLSX para DER_MULTAS.
        """
        try:
            xlsx = load_workbook(self.xlsx_servidor)
            sheet = xlsx.active
            print(f"\nPlanilha DER_MULTAS carregada com sucesso: {sheet.title}")
            return xlsx, sheet
        except Exception as e:
            print(f"Erro ao carregar o arquivo XLSX: {e}")
            return None, None

    def ler_xlsx_DER_Protocolo(self):
        """
        Carrega e exibe informações básicas do arquivo XLSX para DER_Protocolo.
        """
        try:
            xlsx = load_workbook(self.xlsx_servidor)
            sheet = xlsx.active
            print(f"\nPlanilha DER_Protocolo carregada com sucesso: {sheet.title}")
            return xlsx, sheet
        except Exception as e:
            print(f"Erro ao carregar o arquivo XLSX: {e}")
            return None, None
    
    def ler_xlsx_Agentes_Geral(self):
        """
        Carrega e exibe informações básicas do arquivo XLSX para Agentes em Geral
        """
        try:
            xlsx = load_workbook(self.xlsx_servidor)
            sheet = xlsx.active
            print(f"\nPlanilha Agentes_Geral carregada com sucesso: {sheet.title}")
            return xlsx, sheet
        except Exception as e:
            print(f"Erro ao carregar o arquivo XLSX: {e}")
            return None, None

    def modificar_e_salvar_xlsx_DER_MULTAS(self):
        """
        Modifica e salva o arquivo XLSX para a planilha DER_MULTAS.
        """
        # Carregar o arquivo e obter a planilha específica de MULTAS
        xlsx, sheet = self.ler_xlsx_DER_MULTAS()

        if xlsx is not None:
            # Modificar o título da planilha
            print("\nForma Padrão para nomes de planilhas do DER MULTAS: ")
            print("Agent_MêsAno\nExemplo: MULTAS_JANEIRO2025\n ")
            sheet.title = input("Digite o Título da Planilha: ")

            # Definir as alterações de formatação
            cor_fundo = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # Cor cinza
            fonte_negrito = Font(bold=True)  # Negrito
            borda = Border(left=Side(border_style=None), right=Side(border_style=None), 
                           top=Side(border_style=None), bottom=Side(border_style=None))  # Remover bordas

            # Alterar apenas as primeiras linhas das 3 primeiras colunas
            for i in range(1, 4):  # Para as três primeiras colunas
                for row in sheet.iter_rows(min_col=i, max_col=i, max_row=1): 
                    for cell in row:
                        cell.fill = cor_fundo
                        cell.font = fonte_negrito
                        cell.border = borda
            
            # Identificar o número da última linha com dados
            last_row = sheet.max_row

            # Adicionar "TOTAL" em negrito na coluna 2
            total_cell = sheet.cell(row=last_row + 1, column=2)
            total_cell.value = "TOTAL"
            total_cell.font = Font(bold=True)

            # Somar os valores da coluna 3
            total_sum = 0
            for row in range(2, last_row + 1):
                cell_value = sheet.cell(row=row, column=3).value
                if isinstance(cell_value, (int, float)):
                    total_sum += cell_value

            # Adicionar a soma na coluna 3
            sum_cell = sheet.cell(row=last_row + 1, column=3)
            sum_cell.value = total_sum
            
            # Salvar o arquivo de volta no diretório Downloads com um novo nome
            try:
                nome_arquivo = os.path.basename(self.xlsx_servidor)  # Extrai o nome do arquivo
                caminho_salvar = os.path.join(os.environ['USERPROFILE'], 'Downloads', nome_arquivo)
                xlsx.save(caminho_salvar)
                print(f"Arquivo de DER_MULTAS salvo com sucesso: {caminho_salvar}")
            except Exception as e:
                print(f"Erro ao salvar o arquivo DER_MULTAS: {e}")

    def modificar_e_salvar_xlsx_DER_Protocolo(self,titulo_aba):
        """
        Modifica e salva o arquivo XLSX para a planilha DER_Protocolo.
        """
        # Carregar o arquivo e obter a planilha específica de Protocolo
        xlsx, sheet = self.ler_xlsx_DER_Protocolo()

        if xlsx is not None:
            # Modificar o título da planilha
            #print("\nForma Padrão para nomes de planilhas do DER Protocolo: ")
            #print("Agent_MêsAno\nExemplo: PROTOCOLO_JANEIRO2025\n ")
            sheet.title = titulo_aba

            # Definir as alterações de formatação
            cor_fundo = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # Cor cinza
            fonte_negrito = Font(bold=True)  # Negrito
            borda = Border(left=Side(border_style=None), right=Side(border_style=None), 
                           top=Side(border_style=None), bottom=Side(border_style=None))  # Remover bordas

            # Alterar apenas as primeiras linhas das 3 primeiras colunas
            for i in range(1, 4):  # Para as três primeiras colunas
                for row in sheet.iter_rows(min_col=i, max_col=i, max_row=1):
                    for cell in row:
                        cell.fill = cor_fundo
                        cell.font = fonte_negrito
                        cell.border = borda

            # Identificar o número da última linha com dados
            last_row = sheet.max_row

            # Adicionar "TOTAL" em negrito na coluna 2
            total_cell = sheet.cell(row=last_row + 1, column=2)
            total_cell.value = "TOTAL"
            total_cell.font = Font(bold=True)

            # Somar os valores da coluna 3
            total_sum = 0
            for row in range(2, last_row + 1):
                cell_value = sheet.cell(row=row, column=3).value
                if isinstance(cell_value, (int, float)):
                    total_sum += cell_value

            # Adicionar a soma na coluna 3
            sum_cell = sheet.cell(row=last_row + 1, column=3)
            sum_cell.value = total_sum

            # Salvar o arquivo de volta no diretório Downloads com um novo nome
            try:
                nome_arquivo = os.path.basename(self.xlsx_servidor)  # Extrai o nome do arquivo
                caminho_salvar = os.path.join(os.environ['USERPROFILE'], 'Downloads', nome_arquivo)
                xlsx.save(caminho_salvar)
                print(f"Arquivo de DER_Protocolo salvo com sucesso: {caminho_salvar}")
            except Exception as e:
                print(f"Erro ao salvar o arquivo DER_Protocolo: {e}")

    def modificar_e_salvar_xlsx_Agentes_Geral(self):
        """
        Modifica e salva o arquivo XLSX para a planilha Agentes_Geral.
        """
        # Carregar o arquivo e obter a planilha específica de Agentes_Geral
        xlsx, sheet = self.ler_xlsx_Agentes_Geral()

        if xlsx is not None:
            # Modificar o título da planilha
            print("\nForma Padrão para nomes de planilhas do Agentes_Geral: ")
            print("Agent_MêsAno\nExemplo: BENECAP_JANEIRO2025\n ")
            sheet.title = input("Digite o Título da Planilha para o Agent: ")

            # Definir as alterações de formatação
            cor_fundo = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # Cor cinza
            fonte_negrito = Font(bold=True)  # Negrito
            borda = Border(left=Side(border_style=None), right=Side(border_style=None), 
                           top=Side(border_style=None), bottom=Side(border_style=None))  # Remover bordas

            # Alterar apenas as primeiras linhas das 3 primeiras colunas
            for i in range(1, 4):  # Para as três primeiras colunas
                for row in sheet.iter_rows(min_col=i, max_col=i, max_row=1):
                    for cell in row:
                        cell.fill = cor_fundo
                        cell.font = fonte_negrito
                        cell.border = borda


            # Identificar o número da última linha com dados
            last_row = sheet.max_row

            # Adicionar "TOTAL" em negrito na coluna 2
            total_cell = sheet.cell(row=last_row + 1, column=2)
            total_cell.value = "TOTAL"
            total_cell.font = Font(bold=True)

            # Somar os valores da coluna 3
            total_sum = 0
            for row in range(2, last_row + 1):
                cell_value = sheet.cell(row=row, column=3).value
                if isinstance(cell_value, (int, float)):
                    total_sum += cell_value

            # Adicionar a soma na coluna 3
            sum_cell = sheet.cell(row=last_row + 1, column=3)
            sum_cell.value = total_sum
    
            # Salvar o arquivo de volta no diretório Downloads com um novo nome
            try:
                nome_arquivo = os.path.basename(self.xlsx_servidor)  # Extrai o nome do arquivo
                caminho_salvar = os.path.join(os.environ['USERPROFILE'], 'Downloads', nome_arquivo)
                xlsx.save(caminho_salvar)
                print(f"Arquivo de Agentes_Geral salvo com sucesso: {caminho_salvar}")
            except Exception as e:
                print(f"Erro ao salvar o arquivo Agentes_Geral: {e}")

